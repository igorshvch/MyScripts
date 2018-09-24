import re
from collections import deque
from sentan.gui.dialogs import ffp, fdp, pmb
from openpyxl import load_workbook

__version__ = '0.1.1'

###Content=====================================================================

COLS = ('E', 'F', 'AH', 'AJ', 'AK', 'AL', 'AM',)
KLASS = {
    'Может быть разворот' : True,
    'Устоялась есть динамика' : True,
    'Устоялась нет динамики' : True,
    'К удалению' : False
}
VALS = {'ЦИТАТА АКТА', 'ДОБОР НЕ НУЖЕН', 'ЦИТАТА НОРМЫ'}


class RowsCollectror():
    def __init__(self, rows, cols=COLS, break_vals=VALS):
        self.rows = range(*rows)
        self.cols = cols
        self.break_vals = break_vals
        self.wb=load_workbook(filename=r'C:\Users\EA-ShevchenkoIS\Desktop\Работа\ПСП-ХИТЫ-на актуализацию 05.02.2017.xlsx')
        self.wsh = 'Лист1'
    
    def open_excel(self):
        pmb('Выберите книгу Excel')
        self.wb = load_workbook(ffp())
    
    def select_sheet(self):
        print(self.wb.sheetnames)
        self.wsh = input('Введите название рабочего листа ======>> ')
    
    def iterate_over_rows(self):
        holder = []
        wsh = self.wb[self.wsh]
        rows = list(self.rows)
        cols = self.cols
        keys = ['par', 'concl', 'klass', 'pos1', 'pos2', 'pos3', 'pos4']
        for i in rows:
            cells = [col+str(i) for col in cols]
            #print(cells)
            vals = {key:(wsh[cell].value) for key,cell in zip(keys, cells)}
            #print(vals)
            if vals['klass'] == 'К удалению':
                par_concl = vals['par']+'#'+vals['concl']+'#'
                holder.append(par_concl+'К удалению')
            else:
                par_concl = vals['par']+'#'+vals['concl']+'#'
                holder.append(par_concl+ (vals['pos1'] if vals['pos1'] else ''))
                holder.append(par_concl+ (vals['pos2'] if vals['pos2'] else ''))
                holder.append(par_concl+ (vals['pos3'] if vals['pos3'] else ''))
                holder.append(par_concl+ (vals['pos4'] if vals['pos4'] else ''))
        return holder



###Testing=====================================================================
if __name__ == '__main__':
    import sys
    try:
        sys.argv[1]
        if sys.argv[1] == '-v':
            print('Module name: {}'.format(sys.argv[0]))
            print('Version info:', __version__)
        elif sys.argv[1] == '-t':
            print('Testing mode!')
            print('Not implemented!')
        else:
            print('Not implemented!')
    except IndexError:
        print('Mode var wasn\'t passed!')